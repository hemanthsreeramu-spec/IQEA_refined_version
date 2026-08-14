"""
Consolidated result reports — one HTML file and one Excel file per run.

Both carry the same columns and the same colour coding (failures red, skips
amber), so a run can be attached to a ticket or mailed on without anyone having
to open the tool.
"""

import io
from datetime import datetime

import pandas as pd

# The report's column order, and where each value comes from in a result row.
REPORT_COLUMNS = [
    "Test Case",
    "Method",
    "Endpoint",
    "Expected Status",
    "Actual Status",
    "Expected Message",
    "Actual Message",
    "Test Status",
    "Reason",
]

FILL = {
    "FAIL": "FFC7CE",   # red
    "SKIP": "FFEB9C",   # amber
    "PASS": "C6EFCE",   # green
}
TEXT = {
    "FAIL": "9C0006",
    "SKIP": "9C6500",
    "PASS": "006100",
}


def _reason(record):
    """Why a row is not a pass. Empty for passing rows."""
    if str(record.get("Result", "")).upper() == "PASS":
        return ""
    parts = [str(record.get(key) or "").strip() for key in ("Validation", "Note")]
    return " | ".join(part for part in parts if part)


def report_rows(results):
    """Flatten runner results into the reporting columns."""
    rows = []
    for record in results or []:
        rows.append({
            "Test Case": record.get("Test Case") or f"{record.get('Method', '')} {record.get('Endpoint', '')}".strip(),
            "Method": record.get("Method", ""),
            "Endpoint": record.get("Endpoint", ""),
            "Expected Status": record.get("Expected Status", ""),
            "Actual Status": record.get("Status", ""),
            "Expected Message": record.get("Expected Response", "") or "",
            "Actual Message": record.get("Actual Message", "") or "",
            "Test Status": record.get("Result", ""),
            "Reason": _reason(record),
        })
    return rows


def summarise(rows):
    total = len(rows)
    counts = {"PASS": 0, "FAIL": 0, "SKIP": 0}
    for row in rows:
        status = str(row.get("Test Status", "")).upper()
        counts[status] = counts.get(status, 0) + 1
    passed = counts.get("PASS", 0)
    return {
        "total": total,
        "passed": passed,
        "failed": counts.get("FAIL", 0),
        "skipped": counts.get("SKIP", 0),
        "other": total - passed - counts.get("FAIL", 0) - counts.get("SKIP", 0),
        "pass_rate": round((passed / total) * 100, 1) if total else 0.0,
    }


def _escape(value):
    return (
        str("" if value is None else value)
        .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def results_to_html(results, title="API Validation Report", generated_at=None):
    """A standalone, self-contained HTML report."""
    rows = report_rows(results)
    stats = summarise(rows)
    stamp = generated_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    body_rows = []
    for row in rows:
        status = str(row["Test Status"]).upper()
        css = status.lower() if status in FILL else "other"
        cells = "".join(
            f'<td class="{"reason" if column == "Reason" else ""}">{_escape(row[column])}</td>'
            for column in REPORT_COLUMNS
        )
        body_rows.append(f'<tr class="{css}">{cells}</tr>')

    headers = "".join(f"<th>{_escape(column)}</th>" for column in REPORT_COLUMNS)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>{_escape(title)}</title>
<style>
  body {{ font-family: Segoe UI, Arial, sans-serif; color: #1f2430; margin: 24px; background: #fff; }}
  h1 {{ font-size: 22px; margin: 0 0 4px; }}
  .stamp {{ color: #6b7280; font-size: 13px; margin-bottom: 18px; }}
  .cards {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }}
  .card {{ border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px 18px; min-width: 110px; }}
  .card .n {{ font-size: 24px; font-weight: 700; }}
  .card .l {{ font-size: 12px; color: #6b7280; text-transform: uppercase; letter-spacing: .04em; }}
  .card.pass .n {{ color: #067647; }}
  .card.fail .n {{ color: #b42318; }}
  .card.skip .n {{ color: #b54708; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
  th, td {{ border: 1px solid #e5e7eb; padding: 8px 10px; text-align: left; vertical-align: top; }}
  th {{ background: #f3f4f6; position: sticky; top: 0; }}
  td.reason {{ max-width: 380px; word-break: break-word; }}
  tr.pass {{ background: #eafaf1; }}
  tr.fail {{ background: #fde8e8; color: #7f1d1d; font-weight: 600; }}
  tr.skip {{ background: #fef7e0; color: #7c4a03; }}
  .legend {{ margin-top: 14px; font-size: 12px; color: #6b7280; }}
</style>
</head>
<body>
  <h1>{_escape(title)}</h1>
  <div class="stamp">Generated {_escape(stamp)}</div>

  <div class="cards">
    <div class="card"><div class="n">{stats['total']}</div><div class="l">Total</div></div>
    <div class="card pass"><div class="n">{stats['passed']}</div><div class="l">Passed</div></div>
    <div class="card fail"><div class="n">{stats['failed']}</div><div class="l">Failed</div></div>
    <div class="card skip"><div class="n">{stats['skipped']}</div><div class="l">Skipped</div></div>
    <div class="card"><div class="n">{stats['pass_rate']}%</div><div class="l">Pass rate</div></div>
  </div>

  <table>
    <thead><tr>{headers}</tr></thead>
    <tbody>
      {"".join(body_rows) if body_rows else '<tr><td colspan="9">No results.</td></tr>'}
    </tbody>
  </table>
  <div class="legend">Red = failed · Amber = skipped because a dependency failed · Green = passed</div>
</body>
</html>"""


def results_to_excel_bytes(results, sheet_name="Results"):
    """The same report as .xlsx, with failed rows filled red."""
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = report_rows(results)
    frame = pd.DataFrame(rows, columns=REPORT_COLUMNS)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="F3F4F6")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        status_index = REPORT_COLUMNS.index("Test Status") + 1
        for excel_row in range(2, len(frame) + 2):
            status = str(worksheet.cell(row=excel_row, column=status_index).value or "").upper()
            if status not in FILL:
                continue
            fill = PatternFill("solid", fgColor=FILL[status])
            font = Font(color=TEXT[status], bold=(status == "FAIL"))
            for column in range(1, len(REPORT_COLUMNS) + 1):
                cell = worksheet.cell(row=excel_row, column=column)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(vertical="top", wrap_text=(REPORT_COLUMNS[column - 1] == "Reason"))

        widths = {"Test Case": 30, "Method": 9, "Endpoint": 52, "Expected Status": 15,
                  "Actual Status": 13, "Expected Message": 28, "Actual Message": 38,
                  "Test Status": 12, "Reason": 60}
        for index, column in enumerate(REPORT_COLUMNS, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = widths[column]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    buffer.seek(0)
    return buffer.getvalue()
