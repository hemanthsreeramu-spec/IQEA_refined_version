"""
IQEA Mobile Test Case Generator
================================
Converts recorded actions into comprehensive test cases for all possible scenarios.

Features:
  - Auto-detects action types (tap, input, swipe, scroll)
  - Generates positive/negative scenarios
  - Adds edge cases and boundary tests
  - Exports to Excel format compatible with IQEA main platform
  - Generates test case descriptions suitable for manual testers

Flow:
  recorded_actions → generate_test_cases() → test_cases (list[dict])
                                          → export_to_excel() → .xlsx file
"""

import json
import re
from datetime import datetime
from typing import Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class MobileTestCaseGenerator:
    """
    Generates comprehensive test cases from recorded mobile actions.

    A single recorded action can generate:
      - 1 Positive scenario (happy path)
      - 1-2 Negative scenarios (invalid input, missing input)
      - 1 Edge case (boundary values, long text)
    """

    def __init__(self, logger=None):
        self.logger = logger
        self.test_cases: list[dict] = []

    def log(self, msg, level="info"):
        if self.logger:
            getattr(self.logger, level, self.logger.info)(msg)
        else:
            print(f"[{level.upper()}] {msg}")

    def generate_test_cases(self, actions: list[dict], app_name: str = "Mobile App",
                          include_edge_cases=True, include_negative=True) -> list[dict]:
        """
        Convert recorded actions into test cases with multiple scenarios.

        Args:
            actions: List of recorded action dicts (from ActionRecorder.to_dict())
            app_name: Application name for test case titles
            include_edge_cases: Add edge case test cases
            include_negative: Add negative test cases

        Returns:
            List of test case dicts with standardized format
        """
        self.test_cases = []

        if not actions:
            self.log("No actions provided", "warn")
            return []

        # Initialize with basic flow test case (positive scenario)
        self._create_positive_scenario(actions, app_name)

        # Generate additional scenarios
        for i, action in enumerate(actions):
            if action.get("type") == "input":
                if include_negative:
                    self._create_negative_scenario_input(action, i, app_name, actions)
                if include_edge_cases:
                    self._create_edge_case_scenario_input(action, i, app_name)

            elif action.get("type") == "tap":
                if include_negative:
                    self._create_negative_scenario_tap(action, i, app_name)

        self.log(f"Generated {len(self.test_cases)} test cases from {len(actions)} action(s)", "ok")
        return self.test_cases

    def _create_positive_scenario(self, actions: list[dict], app_name: str):
        """Create main positive/happy path test case."""
        steps = []
        for i, action in enumerate(actions):
            step = {
                "step_number": i + 1,
                "description": self._describe_action_positive(action),
                "expected_result": self._expected_result_positive(action),
                "data": action.get("value", ""),
            }
            steps.append(step)

        tc = {
            "test_case_id": "TC_001",
            "test_case_name": f"{app_name} - Happy Path Flow",
            "category": "Workflow",
            "type": "Manual",
            "priority": "High",
            "description": "Complete happy path covering all recorded actions in sequence",
            "precondition": f"1. {app_name} app is installed\n2. Device is connected\n3. App is launched",
            "steps": steps,
            "status": "New",
        }
        self.test_cases.append(tc)

    def _create_negative_scenario_input(self, action: dict, action_idx: int, app_name: str, actions: list[dict]):
        """Create negative scenario for input fields."""
        scenarios = [
            {
                "suffix": "Empty Input",
                "description": f"Leave the '{action.get('label', 'field')}' empty and proceed",
                "expected_result": "System should display validation error or prevent progression",
            },
            {
                "suffix": "Invalid Input",
                "description": f"Enter special characters or invalid data in '{action.get('label', 'field')}'",
                "expected_result": "System should reject invalid input with appropriate error message",
            },
            {
                "suffix": "Max Length Exceeded",
                "description": f"Enter text exceeding maximum allowed length in '{action.get('label', 'field')}'",
                "expected_result": "System should either truncate or reject the input",
            },
        ]

        for i, scenario in enumerate(scenarios):
            steps = []
            # Create steps for all actions, but modify the current action with negative scenario
            for step_idx, step_action in enumerate(actions):
                if step_idx == action_idx:
                    step_desc = scenario["description"]
                    step_result = scenario["expected_result"]
                else:
                    step_desc = self._describe_action_positive(step_action)
                    step_result = self._expected_result_positive(step_action)

                steps.append({
                    "step_number": step_idx + 1,
                    "description": step_desc,
                    "expected_result": step_result,
                })

            tc = {
                "test_case_id": f"TC_NEG_{action_idx:03d}_{i+1:02d}",
                "test_case_name": f"{app_name} - Negative: {scenario['suffix']} (Step {action_idx+1})",
                "category": "Negative",
                "type": "Manual",
                "priority": "Medium",
                "description": f"Test error handling - {scenario['suffix']}",
                "precondition": f"1. {app_name} app is installed and launched",
                "steps": steps,
                "status": "New",
            }
            self.test_cases.append(tc)

    def _create_negative_scenario_tap(self, action: dict, action_idx: int, app_name: str):
        """Create negative scenario for button/clickable elements."""
        tc = {
            "test_case_id": f"TC_NEG_TAP_{action_idx:03d}",
            "test_case_name": f"{app_name} - Tap Disabled Element ('{action.get('label', 'Button')}')",
            "category": "Negative",
            "type": "Manual",
            "priority": "Low",
            "description": f"Verify behavior when '{action.get('label')}' is disabled or unavailable",
            "precondition": f"1. {app_name} app is launched\n2. Element '{action.get('label')}' is disabled/hidden",
            "steps": [
                {
                    "step_number": 1,
                    "description": f"Attempt to tap '{action.get('label')}' when it's disabled",
                    "expected_result": "Element should not respond or show disabled state",
                },
            ],
            "status": "New",
        }
        self.test_cases.append(tc)

    def _create_edge_case_scenario_input(self, action: dict, action_idx: int, app_name: str):
        """Create edge case scenario for input fields."""
        scenarios = [
            {
                "suffix": "Whitespace Only",
                "data": "   ",
                "description": f"Enter only whitespace in '{action.get('label', 'field')}'",
                "expected_result": "System should treat as empty or show validation error",
            },
            {
                "suffix": "Special Characters",
                "data": "!@#$%^&*()",
                "description": f"Enter special characters in '{action.get('label', 'field')}'",
                "expected_result": "System should either accept or show appropriate error",
            },
            {
                "suffix": "Numeric Boundary",
                "data": "999999999",
                "description": f"Enter maximum numeric value in '{action.get('label', 'field')}'",
                "expected_result": "System should process the value correctly",
            },
        ]

        for scenario in scenarios[:1]:  # Limit to 1 edge case per input to avoid explosion
            tc = {
                "test_case_id": f"TC_EDGE_{action_idx:03d}",
                "test_case_name": f"{app_name} - Edge Case: {scenario['suffix']} ('{action.get('label')}')",
                "category": "Edge case",
                "type": "Manual",
                "priority": "Low",
                "description": f"Test boundary conditions - {scenario['suffix']}",
                "precondition": f"1. {app_name} app is launched",
                "steps": [
                    {
                        "step_number": 1,
                        "description": scenario["description"],
                        "expected_result": scenario["expected_result"],
                    },
                ],
                "status": "New",
            }
            self.test_cases.append(tc)

    def _describe_action_positive(self, action: dict) -> str:
        """Generate step description for positive scenario."""
        atype = action.get("type", "tap")
        label = action.get("label", "Element")
        value = action.get("value", "")

        if atype == "input":
            return f"Enter '{value}' in the '{label}' field"
        elif atype == "tap":
            return f"Tap the '{label}' button"
        elif atype == "swipe":
            return f"Swipe the screen"
        elif atype == "scroll":
            return f"Scroll down to reveal more content"
        else:
            return f"Perform action on '{label}'"

    def _expected_result_positive(self, action: dict) -> str:
        """Generate expected result for positive scenario."""
        atype = action.get("type", "tap")
        label = action.get("label", "Element")

        if atype == "input":
            return f"Text is entered in '{label}' field successfully"
        elif atype == "tap":
            return f"'{label}' button is tapped, action executed successfully"
        elif atype == "swipe":
            return f"Screen swipes smoothly, new content visible"
        elif atype == "scroll":
            return f"Page scrolls without errors, more content visible"
        else:
            return f"Action on '{label}' completes successfully"

    def flatten_for_excel(self) -> list[dict]:
        """
        Flatten test cases to single-row-per-step format for Excel export.

        Format:
          Test Case Name | Step Number | Step Description | Step Expected Result | Status | Type | Category
        """
        flattened = []
        for tc in self.test_cases:
            for step in tc.get("steps", []):
                row = {
                    "Test Case Name": tc.get("test_case_name", ""),
                    "Test Case ID": tc.get("test_case_id", ""),
                    "Step Number": step.get("step_number", ""),
                    "Test Step Description": step.get("description", ""),
                    "Test Step Expected Result": step.get("expected_result", ""),
                    "Status": tc.get("status", "New"),
                    "Type": tc.get("type", "Manual"),
                    "Category": tc.get("category", ""),
                    "Priority": tc.get("priority", ""),
                    "Precondition": tc.get("precondition", ""),
                }
                flattened.append(row)

        return flattened

    def export_to_excel(self, output_path: str = None, app_name: str = "MobileApp") -> Optional[str]:
        """
        Export test cases to Excel file.

        Must call generate_test_cases() first.

        Returns:
            Path to generated Excel file, or None if export failed
        """
        if not self.test_cases:
            self.log("No test cases to export. Call generate_test_cases() first.", "warn")
            return None

        if output_path is None:
            # Create output directory if it doesn't exist
            output_dir = Path("Test_Cases_collection")
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(output_dir / f"Test_Cases_{app_name}_{timestamp}.xlsx")
        else:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        try:
            flattened = self.flatten_for_excel()

            if PANDAS_AVAILABLE and OPENPYXL_AVAILABLE:
                self._export_with_formatting(flattened, output_path)
            elif PANDAS_AVAILABLE:
                df = pd.DataFrame(flattened)
                df.to_excel(output_path, index=False, sheet_name="Test Cases")
                self.log(f"Exported to {output_path} (basic formatting)", "ok")
            else:
                self._export_json_fallback(output_path)

            path = Path(output_path).resolve()
            self.log(f"Test cases exported: {path}", "ok")
            return str(path)

        except Exception as exc:
            self.log(f"Export error: {exc}", "err")
            return None

    def _export_with_formatting(self, data: list[dict], output_path: str):
        """Export with Excel formatting."""
        df = pd.DataFrame(data)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Test Cases")
            worksheet = writer.sheets["Test Cases"]

            # Formatting
            header_fill = PatternFill(start_color="F47B20", end_color="F47B20", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = thin_border

            worksheet.column_dimensions["A"].width = 30
            worksheet.column_dimensions["B"].width = 15
            worksheet.column_dimensions["C"].width = 12
            worksheet.column_dimensions["D"].width = 40
            worksheet.column_dimensions["E"].width = 40
            worksheet.column_dimensions["F"].width = 12
            worksheet.column_dimensions["G"].width = 15
            worksheet.column_dimensions["H"].width = 15

        self.log(f"Exported with formatting: {output_path}", "ok")

    def _export_json_fallback(self, output_path: str):
        """Fallback export as JSON if Excel libs unavailable."""
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w") as f:
            json.dump(self.test_cases, f, indent=2)
        self.log(f"Excel not available. Exported to JSON: {json_path}", "warn")

    def export_to_dict(self) -> dict:
        """
        Export test cases as dict for passing to script generator.
        """
        return {
            "test_cases": self.test_cases,
            "generated_at": datetime.now().isoformat(),
            "count": len(self.test_cases),
        }
