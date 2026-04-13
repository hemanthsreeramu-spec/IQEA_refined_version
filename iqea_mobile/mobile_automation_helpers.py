"""
IQEA Mobile Automation — Helpers
==================================
Backend utilities for recording, test case generation, and script generation.

Classes:
  LLMClientWrapper    — Azure OpenAI calls (same pattern as Utilities_Xpath.py)
  RecordingExporter   — Save actions to human-readable .txt + screenshots
  ActionFileManager   — List/load recordings, test cases, scripts from disk
  TestCaseBuilder     — Build LLM prompts for test case generation + parse JSON response
  ScriptBuilder       — Build LLM prompts for script generation

Functions:
  save_recording_workflow()   — Save recording: .txt + screenshots + .json metadata
  save_test_cases_to_excel()  — Save test cases as Excel (single sheet or separate files)
"""

import base64
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from dotenv import load_dotenv; load_dotenv()
import openai

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════════
# PATHS
# ═══════════════════════════════════════════════════════════════════════════════

_BASE          = os.path.dirname(__file__)
OUTPUT_DIR     = os.path.join(_BASE, "..", "output")
RECORDINGS_DIR = os.path.join(OUTPUT_DIR, "mobile_recordings")
TEST_CASES_DIR = os.path.join(OUTPUT_DIR, "mobile_test_cases")
SCRIPTS_DIR    = os.path.join(OUTPUT_DIR, "mobile_scripts")

for _d in [RECORDINGS_DIR, TEST_CASES_DIR, SCRIPTS_DIR]:
    os.makedirs(_d, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# LLM CLIENT  (matches pattern in Utilities_Xpath.py)
# ═══════════════════════════════════════════════════════════════════════════════

class LLMClientWrapper:
    """Azure OpenAI wrapper — same pattern used in Utilities_Xpath.py."""
    def __init__(self):
        os.environ["OPENAI_API_KEY"]  = os.getenv("AZURE_OPENAI_API_KEY", "")
        os.environ["OPENAI_API_BASE"] = os.getenv("AZURE_OPENAI_ENDPOINT", "")
        self.client = openai.OpenAI(
            api_key  = os.environ["OPENAI_API_KEY"],
            base_url = os.environ["OPENAI_API_BASE"],
        )
        self.model = "gpt-5-mini"

    def query(self, formatted_summary: str) -> Optional[str]:
        """Send prompt to LLM, return text response or None on error."""
        print("going inside get_queries_from_ai_updated")
        try:
            response = self.client.chat.completions.create(
                model    = self.model,
                messages = [{"role": "user", "content": formatted_summary}],
            )
            print(response)
            return response.choices[0].message.content
        except Exception as e:
            print(f"[ERROR] LLM call failed: {e}")
            return None


# ═══════════════════════════════════════════════════════════════════════════════
# RECORDING EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════

class RecordingExporter:
    """Convert recorded actions to human-readable .txt and save screenshots."""

    @staticmethod
    def export_to_txt(
        actions: List[Dict],
        device: Dict,
        session_id: str,
        filename: str,
    ) -> Optional[str]:
        """Save actions as a formatted .txt file. Returns saved path or None."""
        try:
            filepath = os.path.join(RECORDINGS_DIR, f"{filename}.txt")
            lines = [
                "=" * 80,
                "IQEA MOBILE AUTOMATION — RECORDED SESSION",
                "=" * 80,
                f"Generated  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Device     : {device.get('model', 'Unknown')} ({device.get('platform', 'Unknown')})",
                f"OS         : {device.get('platform', '')} {device.get('os', '')}".strip(),
                f"Session ID : {session_id}",
                f"Total Steps: {len(actions)}",
                "",
                "=" * 80,
                "RECORDED ACTIONS",
                "=" * 80,
            ]

            for i, action in enumerate(actions, 1):
                lines += [
                    "",
                    f"[STEP #{i:02d}]  {action.get('type', 'unknown').upper()}",
                    f"  Time      : {action.get('ts', 'N/A')}",
                    f"  Label     : {action.get('label', 'N/A')}",
                    f"  XPath     : {action.get('xpath', 'N/A')}",
                ]
                if action.get("value"):
                    lines.append(f"  Value     : {action['value']}")
                if action.get("start"):
                    lines.append(f"  From      : {action['start']}")
                if action.get("end"):
                    lines.append(f"  To        : {action['end']}")
                lines.append("-" * 40)

            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            return filepath
        except Exception as e:
            print(f"[ERROR] export_to_txt failed: {e}")
            return None

    @staticmethod
    def save_screenshots(actions: List[Dict], filename: str) -> str:
        """Decode base64 screenshots from actions and save to a folder. Returns folder path."""
        try:
            ss_dir = os.path.join(RECORDINGS_DIR, f"{filename}_screenshots")
            os.makedirs(ss_dir, exist_ok=True)

            for i, action in enumerate(actions):
                if action.get("screenshot_b64"):
                    img_data = base64.b64decode(action["screenshot_b64"])
                    img_path = os.path.join(
                        ss_dir,
                        f"step_{i + 1:03d}_{action.get('type', 'unknown')}.png",
                    )
                    with open(img_path, "wb") as f:
                        f.write(img_data)

            return ss_dir
        except Exception as e:
            print(f"[ERROR] save_screenshots failed: {e}")
            return ""


# ═══════════════════════════════════════════════════════════════════════════════
# ACTION FILE MANAGER
# ═══════════════════════════════════════════════════════════════════════════════

class ActionFileManager:
    """List and load recording, test case, and script files from disk."""

    # ── Recordings ────────────────────────────────────────────────────────────

    @staticmethod
    def list_recording_files() -> List[str]:
        """Return sorted list of .txt recording filenames in RECORDINGS_DIR (user-friendly)."""
        try:
            return sorted(
                f for f in os.listdir(RECORDINGS_DIR)
                if f.endswith(".txt")
            )
        except Exception:
            return []

    @staticmethod
    def _recording_json_path(filename: str) -> str:
        """Resolve the backing JSON path from a .txt or _recordings.json filename."""
        base = filename.replace(".txt", "").replace("_recordings.json", "")
        return os.path.join(RECORDINGS_DIR, f"{base}_recordings.json")

    @staticmethod
    def load_recording(filename: str) -> Optional[Dict]:
        """Load recording data. Accepts .txt filename (shown in UI) or _recordings.json."""
        try:
            path = ActionFileManager._recording_json_path(filename)
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            return None
        except Exception as e:
            print(f"[ERROR] load_recording {filename}: {e}")
            return None

    @staticmethod
    def get_screenshot_paths(recording_filename: str) -> List[str]:
        """Return sorted list of .png paths for a recording. Accepts .txt filename."""
        base   = recording_filename.replace(".txt", "").replace("_recordings.json", "")
        ss_dir = os.path.join(RECORDINGS_DIR, f"{base}_screenshots")
        if os.path.exists(ss_dir):
            return sorted(
                os.path.join(ss_dir, f)
                for f in os.listdir(ss_dir)
                if f.endswith(".png")
            )
        return []

    # ── Test Cases ────────────────────────────────────────────────────────────

    @staticmethod
    def list_test_case_files() -> List[str]:
        """Return sorted list of .xlsx test case filenames in TEST_CASES_DIR (user-friendly)."""
        try:
            return sorted(
                f for f in os.listdir(TEST_CASES_DIR)
                if f.endswith(".xlsx")
            )
        except Exception:
            return []

    @staticmethod
    def _testcase_json_path(filename: str) -> str:
        """Resolve the backing JSON path from a .xlsx or _testcases.json filename."""
        base = filename.replace(".xlsx", "").replace("_testcases.json", "")
        return os.path.join(TEST_CASES_DIR, f"{base}_testcases.json")

    @staticmethod
    def load_test_cases(filename: str) -> List[Dict]:
        """
        Load test case data. Accepts .xlsx filename (shown in UI) or _testcases.json.

        Strategy:
          1. Try exact backing JSON  → {base}_testcases.json
          2. For separate-mode files like base_TC_001.xlsx, strip the _TC_NNN suffix
             and try the parent JSON  → {root}_testcases.json
          3. Fall back to reading the .xlsx directly row by row
        """
        try:
            # ── 1. Try exact JSON ───────────────────────────────────────────────
            json_path = ActionFileManager._testcase_json_path(filename)
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    return json.load(f)

            # ── 2. Strip _TC_NNN / _TC-NNN suffix and try parent JSON ──────────
            base = filename.replace(".xlsx", "").replace("_testcases.json", "")
            parent_base = re.sub(r"_TC[_-]?\d+$", "", base, flags=re.IGNORECASE)
            if parent_base != base:
                parent_json = os.path.join(TEST_CASES_DIR, f"{parent_base}_testcases.json")
                if os.path.exists(parent_json):
                    with open(parent_json, "r", encoding="utf-8") as f:
                        return json.load(f)

            # ── 3. Fall back: read .xlsx directly ──────────────────────────────
            xlsx_path = os.path.join(TEST_CASES_DIR, filename)
            if xlsx_path.endswith(".xlsx") and os.path.exists(xlsx_path) and HAS_OPENPYXL:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
                test_cases = []
                for ws in wb.worksheets:
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        continue
                    headers = [str(h).strip() if h else "" for h in rows[0]]
                    for row in rows[1:]:
                        if not any(row):
                            continue
                        tc = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
                        test_cases.append(tc)
                wb.close()
                return test_cases

            print(f"[WARN] load_test_cases: no data found for '{filename}'")
            return []

        except Exception as e:
            print(f"[ERROR] load_test_cases {filename}: {e}")
            return []

    # ── Scripts ───────────────────────────────────────────────────────────────

    @staticmethod
    def list_script_files() -> List[str]:
        """Return sorted list of filenames in SCRIPTS_DIR."""
        try:
            return sorted(os.listdir(SCRIPTS_DIR))
        except Exception:
            return []


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER FOR TEST CASES
# ═══════════════════════════════════════════════════════════════════════════════

_TC_COLUMNS = [
    "TC ID", "TC Name", "Category", "Type", "Priority",
    "Description", "Precondition", "Step No", "Test Step",
    "Expected Result", "Actual Result", "Status",
]
_COL_WIDTHS = [10, 32, 14, 16, 10, 36, 30, 8, 42, 42, 25, 12]
_ORANGE = "F47B20"
_WHITE  = "FFFFFF"
_LIGHT  = "FFF3E0"


def _apply_header(ws) -> None:
    """Apply styled orange header row to an openpyxl worksheet."""
    hdr_fill   = PatternFill("solid", fgColor=_ORANGE)
    hdr_font   = Font(bold=True, color=_WHITE, size=10)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, col_name in enumerate(_TC_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border

    ws.row_dimensions[1].height = 30
    for col_i, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = width


def _write_tc(ws, tc: Dict, start_row: int, even: bool) -> int:
    """Write one test case (all its steps) into the worksheet. Returns next free row."""
    steps = tc.get("steps") or [
        {"step_no": 1, "step_desc": tc.get("description", ""), "expected": tc.get("expected_result", "")}
    ]
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap   = Alignment(vertical="top", wrap_text=True)
    fill   = PatternFill("solid", fgColor=_LIGHT if even else _WHITE)

    for step_i, step in enumerate(steps):
        row = start_row + step_i
        row_vals = [
            tc.get("tc_id",        ""),
            tc.get("tc_name",      ""),
            tc.get("category",     ""),
            tc.get("type",         ""),
            tc.get("priority",     ""),
            tc.get("description",  "") if step_i == 0 else "",
            tc.get("precondition", "") if step_i == 0 else "",
            step.get("step_no",    step_i + 1),
            step.get("step_desc",  step.get("description", "")),
            step.get("expected",   step.get("expected_result", "")),
            "",  # Actual Result — filled by tester
            "",  # Status        — filled by tester
        ]
        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = wrap

    return start_row + len(steps)


def save_test_cases_to_excel(
    test_cases: List[Dict],
    filename: str,
    mode: str = "single",       # "single" | "separate"
) -> List[str]:
    """
    Persist test cases to Excel.

    Args:
        test_cases : list of TC dicts from LLM / parse_json_response()
        filename   : base name (no extension) used for output files
        mode       : "single"   → one .xlsx, all TCs in one sheet
                     "separate" → one .xlsx per TC

    Returns:
        List of absolute paths to saved Excel files.

    Side-effect:
        Always saves {filename}_testcases.json for Tab 3 to load later.
    """
    if not HAS_OPENPYXL:
        print("[ERROR] openpyxl is not installed — cannot export to Excel")
        return []

    saved = []

    # ── always persist master JSON so Tab 3 can load these TCs ──────────────
    json_path = os.path.join(TEST_CASES_DIR, f"{filename}_testcases.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(test_cases, f, indent=2, ensure_ascii=False)

    if mode == "single":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        ws.freeze_panes = "A2"
        _apply_header(ws)

        current_row = 2
        for tc_i, tc in enumerate(test_cases):
            current_row = _write_tc(ws, tc, current_row, even=(tc_i % 2 == 0))

        out = os.path.join(TEST_CASES_DIR, f"{filename}.xlsx")
        wb.save(out)
        saved.append(out)

    else:  # separate
        for tc_i, tc in enumerate(test_cases):
            tc_id   = re.sub(r'[\\/:*?"<>|]', "_", tc.get("tc_id",   f"TC_{tc_i+1:03d}"))
            tc_name = re.sub(r'[\\/:*?"<>|]', "_", tc.get("tc_name", "test_case"))[:30]
            fname   = f"{filename}_{tc_id}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = tc_id[:31]
            ws.freeze_panes = "A2"
            _apply_header(ws)
            _write_tc(ws, tc, 2, even=False)

            out = os.path.join(TEST_CASES_DIR, fname)
            wb.save(out)
            saved.append(out)

    return saved


# ═══════════════════════════════════════════════════════════════════════════════
# TEST CASE BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

class TestCaseBuilder:
    """Build LLM prompts for test case generation and parse the JSON response."""

    @staticmethod
    def extract_text_from_screenshots(
        screenshot_paths: List[str],
        selected_indices: List[int],
        actions: List[Dict],
    ) -> str:
        """
        Run Tesseract OCR on selected screenshots and return combined extracted text.
        Falls back to base64-embedded screenshots in action dicts if file not found.
        Returns empty string if pytesseract is not available.
        """
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            return ""

        extracted_parts = []
        for idx in selected_indices:
            img = None
            label = f"Step {idx + 1}"
            # Try file path first
            if idx < len(screenshot_paths):
                try:
                    img = Image.open(screenshot_paths[idx])
                    label = f"Step {idx + 1} ({Path(screenshot_paths[idx]).name})"
                except Exception:
                    img = None
            # Fallback to base64 in action dict
            if img is None and idx < len(actions) and actions[idx].get("screenshot_b64"):
                try:
                    import base64
                    img_bytes = base64.b64decode(actions[idx]["screenshot_b64"])
                    img = Image.open(io.BytesIO(img_bytes))
                except Exception:
                    img = None

            if img:
                try:
                    text = pytesseract.image_to_string(img).strip()
                    if text:
                        extracted_parts.append(
                            f"\n--- {label} ---\n{text}"
                        )
                    else:
                        extracted_parts.append(
                            f"\n--- {label} ---\n(no text detected)"
                        )
                except Exception as e:
                    extracted_parts.append(
                        f"\n--- {label} ---\n(OCR error: {e})"
                    )

        return "\n".join(extracted_parts)

    @staticmethod
    def build_prompt(
        actions: List[Dict],
        selected_screenshot_indices: List[int],
        user_description: str,
        app_name: str = "Mobile App",
        image_extracted_text: str = "",
    ) -> str:
        """
        Return a prompt string asking the LLM for test cases in JSON format.

        Args:
            image_extracted_text: Text extracted from screenshots via Tesseract OCR.
                                  If empty, prompt is action-only (no screenshot context).
        """

        action_lines = []
        for i, a in enumerate(actions):
            line = f"  {i+1}. [{a.get('type','?').upper()}] {a.get('label','N/A')}"
            if a.get("value"):
                line += f"  →  entered: '{a['value']}'"
            if a.get("xpath"):
                line += f"\n       XPath: {a['xpath']}"
            action_lines.append(line)

        # Build screenshot section — rich if OCR data present, brief if not
        if image_extracted_text.strip():
            screenshot_section = f"""
SCREENSHOT UI CONTENT (extracted via OCR from {len(selected_screenshot_indices)} selected screenshot(s)):
{image_extracted_text}

Use the above UI text to understand exactly what labels, buttons, input fields, messages, and
navigation elements are visible on each screen. Reference these in your test case steps and
expected results for accuracy.
"""
        elif selected_screenshot_indices:
            screenshot_section = (
                f"\nUser selected screenshots at steps: "
                f"{', '.join(str(i+1) for i in selected_screenshot_indices)}\n"
            )
        else:
            screenshot_section = "\n(No screenshots used — generating from recorded actions only)\n"

        return f"""You are a senior QA engineer. Generate comprehensive test cases for a mobile application.

APPLICATION: {app_name}

RECORDED USER ACTIONS:
{chr(10).join(action_lines)}
{screenshot_section}
USER REQUIREMENTS / FOCUS AREAS:
{user_description}

Generate test cases covering Positive, Negative, and Edge Case scenarios.
Where screenshot UI content is provided, use the exact button labels, field names, and
messages visible in the UI for accurate test steps and expected results.

Return ONLY a valid JSON array — no markdown fences, no explanations, nothing else.
Use this exact structure:

[
  {{
    "tc_id": "TC_001",
    "tc_name": "Concise test case name",
    "category": "Positive",
    "type": "Functional",
    "priority": "High",
    "description": "What this test case validates",
    "precondition": "App state / data required before test",
    "steps": [
      {{"step_no": 1, "step_desc": "Action to perform", "expected": "Expected outcome"}},
      {{"step_no": 2, "step_desc": "Next action",       "expected": "Expected outcome"}}
    ]
  }}
]

Allowed values:
  category : Positive | Negative | Edge Case | UI | Performance
  type     : Functional | Non-Functional | Security | Usability | Compatibility
  priority : Critical | High | Medium | Low

Generate at least 10 diverse test cases. Return ONLY the JSON array."""

    @staticmethod
    def parse_json_response(response_text: str) -> List[Dict]:
        """Extract and parse the JSON array of test cases from LLM response."""
        if not response_text:
            return []

        # Try direct parse first (LLM returned pure JSON)
        try:
            data = json.loads(response_text.strip())
            if isinstance(data, list):
                return data
        except Exception:
            pass

        # Try to extract [...] block
        try:
            match = re.search(r"\[[\s\S]*\]", response_text)
            if match:
                data = json.loads(match.group())
                if isinstance(data, list):
                    return data
        except Exception:
            pass

        print("[WARN] Could not parse test case JSON from LLM response")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# SCRIPT BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

class ScriptBuilder:
    """Build LLM prompts for executable script generation."""

    LANGUAGE_MAP: Dict[str, Dict[str, str]] = {
        "Python":          {"ext": "py",    "framework": "Appium (pytest)"},
        "Java":            {"ext": "java",  "framework": "TestNG + Appium"},
        "JavaScript":      {"ext": "js",    "framework": "WebdriverIO"},
        "C#":              {"ext": "cs",    "framework": "NUnit + Appium"},
        "C++":             {"ext": "cpp",   "framework": "Appium C++"},
        "Ruby":            {"ext": "rb",    "framework": "RSpec + Appium"},
        "Robot Framework": {"ext": "robot", "framework": "AppiumLibrary"},
    }

    @staticmethod
    def build_prompt(
        test_cases: List[Dict],
        actions: List[Dict],
        language: str,
        custom_requirements: str = "",
    ) -> str:
        """Return a prompt asking LLM to generate a complete executable test script."""

        framework = ScriptBuilder.LANGUAGE_MAP.get(language, {}).get("framework", "Appium")

        tc_lines = "\n".join(
            f"  {tc.get('tc_id','')}: {tc.get('tc_name','')} "
            f"[{tc.get('category','')} | {tc.get('priority','')}]"
            for tc in test_cases
        )

        action_lines = "\n".join(
            "  {seq}. [{t}] {lbl}{xpath}{val}".format(
                seq  = i + 1,
                t    = a.get("type", "?").upper(),
                lbl  = a.get("label", "N/A"),
                xpath = f"\n       xpath='{a['xpath']}'" if a.get("xpath") else "",
                val   = f"  value='{a['value']}'"        if a.get("value") else "",
            )
            for i, a in enumerate(actions)
        )

        extra = f"\nADDITIONAL REQUIREMENTS:\n{custom_requirements}\n" if custom_requirements.strip() else ""

        appium2_setup = ""
        if language == "Python":
            appium2_setup = """
CRITICAL — APPIUM 2.x / 3.x SETUP (Python):
- Server URL MUST be "http://127.0.0.1:4723"  (use 127.0.0.1 not localhost, NO /wd/hub suffix)
- Use UiAutomator2Options class, NOT a raw DesiredCapabilities dict
- Imports:
    from appium import webdriver
    from appium.options import UiAutomator2Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
- Build driver (exact pattern — do not deviate):
    options = UiAutomator2Options()
    options.platform_name        = "Android"
    options.platform_version     = "13"                    # update to match device OS version
    options.device_name          = "YOUR_DEVICE_NAME"      # update e.g. "Samsung SM-A226B"
    options.udid                 = "YOUR_DEVICE_UDID"      # update e.g. "R9ZR701FW6R"
    options.automation_name      = "UiAutomator2"
    options.app_package          = "com.example.app"       # update to real app package
    options.app_activity         = ".MainActivity"         # update to real main activity
    options.no_reset             = True
    options.new_command_timeout  = 300
    driver = webdriver.Remote("http://127.0.0.1:4723", options=options)
- pytest fixture scope MUST be "module" with yield and driver.quit() in teardown
- Define ALL XPath locators as NAMED CONSTANTS at the top of the file
- Add helper functions: wait_for_element(), tap_element(), input_text(), is_element_present()
- Use pytest.skip() when optional elements are not present — do NOT use pytest.fail() for optional UI
- Include allure decorators: @allure.feature(), @allure.story(), with allure.step()
"""
        elif language == "Java":
            appium2_setup = """
CRITICAL — APPIUM 2.x / 3.x SETUP (Java):
- Server URL: new URL("http://127.0.0.1:4723")  (use 127.0.0.1, NO /wd/hub)
- Use UiAutomator2Options, NOT DesiredCapabilities
- Driver: new AndroidDriver(new URL("http://127.0.0.1:4723"), options)
- Include UDID: options.setUdid("YOUR_DEVICE_UDID")
"""
        elif language == "JavaScript":
            appium2_setup = """
CRITICAL — APPIUM 2.x / 3.x SETUP (WebdriverIO):
- hostname: '127.0.0.1', port: 4723, path: '/'
- capabilities: [{ 'appium:automationName': 'UiAutomator2', 'appium:udid': 'YOUR_UDID', ... }]
"""

        return f"""You are an expert mobile automation engineer.
Generate a complete, executable {language} test automation script.

LANGUAGE  : {language}
FRAMEWORK : {framework}
{appium2_setup}
TEST CASES TO IMPLEMENT:
{tc_lines}

RECORDED ACTIONS WITH EXACT XPATH LOCATORS:
{action_lines}
{extra}
Script requirements:
1. Implement ALL test cases listed above as individual test methods/functions with docstrings
2. Use {framework} for mobile automation
3. Define ALL XPath locators as named constants at the top of the file
4. Use Appium 2.x setup as shown above — NO DesiredCapabilities dict, NO /wd/hub URL
5. Add helper functions: wait_for_element, tap_element, input_text, is_element_present
6. Add meaningful assertions that verify each expected outcome
7. Use pytest.skip() (or equivalent) when optional elements are not present rather than failing
8. Include proper error handling, explicit waits (WebDriverWait), and logging
9. Follow {language} naming conventions and best practices
10. Add inline comments explaining each step
11. The script must be ready to run — only device caps (platformVersion, deviceName, app details) need updating

Return ONLY the executable {language} code. No markdown fences. No explanations."""

    @staticmethod
    def get_extension(language: str) -> str:
        return ScriptBuilder.LANGUAGE_MAP.get(language, {}).get("ext", "txt")


# ═══════════════════════════════════════════════════════════════════════════════
# SAVE RECORDING WORKFLOW  (used by Tab 2)
# ═══════════════════════════════════════════════════════════════════════════════

def save_recording_workflow(
    actions: List[Dict],
    device: Dict,
    session_id: str,
    filename: str,
) -> Dict[str, Any]:
    """
    Save a completed recording:
      • {filename}.txt            — human-readable action list
      • {filename}_screenshots/  — one PNG per recorded step
      • {filename}_recordings.json — full JSON metadata (loaded by Tab 2B / Tab 3)

    Returns dict: {txt_file, json_file, screenshots_dir, error}
    """
    result: Dict[str, Any] = {
        "txt_file":       None,
        "json_file":      None,
        "screenshots_dir": None,
        "error":          None,
    }

    try:
        result["txt_file"]       = RecordingExporter.export_to_txt(actions, device, session_id, filename)
        result["screenshots_dir"] = RecordingExporter.save_screenshots(actions, filename)

        json_path = os.path.join(RECORDINGS_DIR, f"{filename}_recordings.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "filename":   filename,
                    "actions":    actions,
                    "device":     device,
                    "session_id": session_id,
                    "created_at": datetime.now().isoformat(),
                },
                f,
                indent=2,
                ensure_ascii=False,
            )
        result["json_file"] = json_path

    except Exception as e:
        result["error"] = str(e)
        print(f"[ERROR] save_recording_workflow: {e}")

    return result
