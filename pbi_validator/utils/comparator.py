import re
from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Value normalisation ────────────────────────────────────────────────────────

def normalize_value(raw):
    """
    Convert a display string to a float so two values can be compared numerically.

    Examples:
        "$1.2M"   → 1_200_000.0
        "94.3%"   → 94.3
        "1,234"   → 1_234.0
        "12.5K"   → 12_500.0
        "3.1B"    → 3_100_000_000.0
        "N/A"     → None
    """
    if raw is None:
        return None

    s = str(raw).strip()

    # Remove currency symbols
    s = re.sub(r"[$€£¥₹]", "", s)
    # Remove commas and spaces
    s = s.replace(",", "").replace(" ", "")

    is_pct = s.endswith("%")
    s = s.rstrip("%")

    multiplier = 1
    upper = s.upper()
    if upper.endswith("B"):
        multiplier = 1_000_000_000
        s = s[:-1]
    elif upper.endswith("M"):
        multiplier = 1_000_000
        s = s[:-1]
    elif upper.endswith("K"):
        multiplier = 1_000
        s = s[:-1]

    try:
        return float(s) * multiplier
    except (ValueError, TypeError):
        return None


# ── Comparison ─────────────────────────────────────────────────────────────────

def compare_values(ui_value: str, db_value, tolerance_pct: float = 0.1) -> tuple:
    """
    Compare a UI display value against the DB result.

    Returns (status, reason) where status is "PASS", "FAIL", or "ERROR".
    tolerance_pct is the acceptable percentage difference (default 0.1 %).
    """
    if db_value is None:
        return "FAIL", "DB query returned no result"

    ui_norm = normalize_value(ui_value)
    db_norm = normalize_value(db_value)

    # ── String comparison when neither side parses as a number ──────────────
    if ui_norm is None and db_norm is None:
        if str(ui_value).strip().lower() == str(db_value).strip().lower():
            return "PASS", "Exact text match"
        return "FAIL", f"Text mismatch — UI: '{ui_value}'  DB: '{db_value}'"

    if ui_norm is None:
        return "ERROR", f"Cannot parse UI value '{ui_value}' as a number"

    if db_norm is None:
        return "ERROR", f"Cannot parse DB value '{db_value}' as a number"

    # ── Percentage scale reconciliation ─────────────────────────────────────
    # PBI shows a percent (e.g. "6.96 %") while SQL often returns the raw
    # fraction (0.0696). Scale the fraction up so they compare equal.
    ui_is_pct = str(ui_value).strip().endswith("%")
    db_is_pct = str(db_value).strip().endswith("%")
    if ui_is_pct and not db_is_pct and abs(db_norm) <= 1.5:
        db_norm *= 100

    # ── Both zero ────────────────────────────────────────────────────────────
    if ui_norm == 0 and db_norm == 0:
        return "PASS", "Both values are zero"

    if ui_norm == 0:
        return "FAIL", f"UI value is 0 but DB returned {db_norm}"

    diff_pct = abs(ui_norm - db_norm) / abs(ui_norm) * 100

    if diff_pct <= tolerance_pct:
        return "PASS", f"Within {tolerance_pct}% tolerance (diff = {diff_pct:.4f}%)"

    return (
        "FAIL",
        f"Diff {diff_pct:.2f}% exceeds {tolerance_pct}% tolerance  "
        f"(UI ≈ {ui_norm:,.2f}  |  DB ≈ {db_norm:,.2f})"
    )


# ── Visual-group comparison ────────────────────────────────────────────────────

def compare_visual_result(visual_entry: dict, db_result,
                           tolerance_pct: float = 0.1) -> list:
    """
    Compare all KPIs in a visual group against the DB query result.

    db_result may be:
      - None          → query returned no rows
      - scalar        → single-KPI scalar (execute_query shortcut)
      - pd.DataFrame  → full result set

    Returns a flat list of per-KPI result dicts (same shape expected by
    build_validation_df / the validation results display).
    """
    visual_name = visual_entry.get("visual_name", "")
    visual_type = visual_entry.get("visual_type", "card")
    sql_query   = visual_entry.get("sql_query", "")
    kpis        = visual_entry.get("kpis", [])

    def _make(kpi, db_val, status, reason):
        return {
            "visual_name": visual_name,
            "kpi_name":    kpi.get("kpi_name", ""),
            "ui_value":    kpi.get("ui_value", ""),
            "sql_query":   sql_query,
            "db_value":    str(db_val) if db_val is not None else "",
            "status":      status,
            "reason":      reason,
        }

    # ── No result ─────────────────────────────────────────────────────────────
    is_empty = (
        db_result is None
        or (isinstance(db_result, pd.DataFrame) and db_result.empty)
    )
    if is_empty:
        return [_make(k, None, "FAIL", "DB query returned no rows") for k in kpis]

    # ── Scalar shortcut (execute_query returns scalar when shape=(1,1)) ───────
    if not isinstance(db_result, pd.DataFrame):
        if len(kpis) == 1:
            s, r = compare_values(kpis[0]["ui_value"], db_result, tolerance_pct)
            return [_make(kpis[0], db_result, s, r)]
        return [_make(k, db_result, "ERROR",
                      "Expected multi-column DataFrame but got a scalar") for k in kpis]

    # ── DataFrame result ──────────────────────────────────────────────────────
    is_card = visual_type in ("card", "kpi_visual", "metric", "other")
    col_lower = {c.lower(): c for c in db_result.columns}  # case-insensitive lookup

    results = []
    for kpi in kpis:
        col_alias     = (kpi.get("column_alias") or "").strip()
        row_key_col   = kpi.get("row_key_column")
        row_key_val   = kpi.get("row_key_value")

        # Resolve target column — exact match first, then case-insensitive, then last column
        if col_alias and col_alias in db_result.columns:
            target_col = col_alias
        elif col_alias and col_alias.lower() in col_lower:
            target_col = col_lower[col_alias.lower()]
        elif db_result.columns.tolist():
            target_col = db_result.columns[-1]
        else:
            results.append(_make(kpi, None, "ERROR",
                                 f"Column '{col_alias}' not found in query result"))
            continue

        if is_card or not row_key_col:
            # Card / single-row — compare first row of target column
            db_val = db_result[target_col].iloc[0]
            s, r   = compare_values(kpi["ui_value"], db_val, tolerance_pct)
            results.append(_make(kpi, db_val, s, r))
        else:
            # Table / chart — find the row where row_key_column = row_key_value
            rk_col_actual = col_lower.get(str(row_key_col).lower(), row_key_col)
            if rk_col_actual not in db_result.columns:
                results.append(_make(kpi, None, "ERROR",
                                     f"Row-key column '{row_key_col}' not found in result"))
                continue

            mask     = (db_result[rk_col_actual].astype(str).str.strip().str.lower()
                        == str(row_key_val).strip().lower())
            matching = db_result[mask]

            if matching.empty:
                results.append(_make(kpi, None, "FAIL",
                                     f"No row found where {row_key_col} = '{row_key_val}'"))
            else:
                db_val = matching[target_col].iloc[0]
                s, r   = compare_values(kpi["ui_value"], db_val, tolerance_pct)
                results.append(_make(kpi, db_val, s, r))

    return results


# ── Phase 3 helpers ────────────────────────────────────────────────────────────

def parse_kpi_name(kpi_name: str) -> tuple:
    """
    Parse 'Total Sales [North Asia]' → ('Total Sales', 'North Asia')
    Parse 'Total Sales'              → ('Total Sales', None)
    """
    m = re.match(r'^(.+?)\s*\[(.+)\]\s*$', kpi_name.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return kpi_name.strip(), None


def match_kpi_to_query(metric: str, reference_queries: list) -> tuple:
    """
    Rule-based: find the reference query whose kpi_patterns best match metric.
    Returns (query_dict, column_alias) or (None, None).
    """
    metric_l = metric.lower().strip()

    for query in reference_queries:
        patterns = query.get("kpi_patterns", [])
        columns  = query.get("metric_columns", [])
        for i, pat in enumerate(patterns):
            if pat.lower().strip() == metric_l:
                col = columns[i] if i < len(columns) else (columns[0] if columns else None)
                return query, col

    # Partial / fuzzy pass
    for query in reference_queries:
        patterns = query.get("kpi_patterns", [])
        columns  = query.get("metric_columns", [])
        for i, pat in enumerate(patterns):
            pat_l = pat.lower().strip()
            if pat_l in metric_l or metric_l in pat_l:
                col = columns[i] if i < len(columns) else (columns[0] if columns else None)
                return query, col

    return None, None


def match_kpi_to_queries(metric: str, reference_queries: list) -> list:
    """
    Return ALL (query, column) candidates whose kpi_patterns match `metric`
    (exact matches first, then fuzzy). A metric like 'Total Sales' appears in
    several visuals with different dimensions — returning every candidate lets
    the caller disambiguate by which query's result actually holds the KPI's
    dimension value.
    """
    metric_l = metric.lower().strip()
    exact, fuzzy = [], []
    for query in reference_queries:
        patterns = query.get("kpi_patterns", [])
        columns  = query.get("metric_columns", [])
        for i, pat in enumerate(patterns):
            pat_l = pat.lower().strip()
            col   = columns[i] if i < len(columns) else (columns[0] if columns else None)
            if pat_l == metric_l:
                exact.append((query, col))
            elif pat_l and (pat_l in metric_l or metric_l in pat_l):
                fuzzy.append((query, col))

    seen, out = set(), []
    for q, c in exact + fuzzy:
        if id(q) not in seen:
            seen.add(id(q))
            out.append((q, c))
    return out


def _cell_matches(part, cell) -> bool:
    """
    Flexible match of one dimension part against a DB cell value:
      - exact (case/space-insensitive), else
      - if BOTH contain digits, compare digits only ("Qtr 1" ↔ 1, "2011" ↔ 2011), else
      - substring either way for text dimensions ("APAC" ↔ "APAC").
    """
    def _norm(x):
        s = str(x).strip().lower()
        if re.fullmatch(r"-?\d+\.0+", s):   # pandas upcasts int rows to float: 2011.0 → 2011
            s = s.split(".")[0]
        return s

    p = _norm(part)
    c = _norm(cell)
    if not p or not c:
        return False
    if p == c:
        return True
    p_dig = re.sub(r"\D", "", p)
    c_dig = re.sub(r"\D", "", c)
    if p_dig and c_dig:
        return p_dig == c_dig
    return p in c or c in p


def extract_db_value(metric: str, dim_value, query: dict, db_result, col_alias: str = None):
    """
    Pull the relevant cell from db_result for this KPI.
    db_result may be None, a scalar, or a DataFrame.
    """
    if db_result is None:
        return None

    if not isinstance(db_result, pd.DataFrame):
        return db_result  # scalar short-circuit

    if db_result.empty:
        return None

    col_lower = {c.lower(): c for c in db_result.columns}

    # Resolve target column
    target_col = None
    if col_alias:
        if col_alias in db_result.columns:
            target_col = col_alias
        elif col_alias.lower() in col_lower:
            target_col = col_lower[col_alias.lower()]

    if not target_col:
        # Fuzzy match metric words against column names
        metric_words = set(re.sub(r"[^\w]", " ", metric).lower().split())
        for col in db_result.columns:
            col_words = set(re.sub(r"[^\w]", " ", col).lower().split())
            if metric_words & col_words:
                target_col = col
                break

    if not target_col:
        # Last resort: last numeric column
        for col in reversed(db_result.columns.tolist()):
            if pd.api.types.is_numeric_dtype(db_result[col]):
                target_col = col
                break

    if not target_col:
        return None

    # Card (no dim_value) → first row
    if not dim_value:
        return db_result[target_col].iloc[0]

    # Table → find the row whose dimension column(s) jointly match dim_value.
    # dim_value may be compound, e.g. "2011, Qtr 1" for a Year+Quarter visual,
    # so split into parts and require EVERY part to match some non-metric column
    # in the same row (digit-aware, so "Qtr 1" matches a Quarter value of 1).
    parts    = [p.strip() for p in str(dim_value).split(",") if p.strip()]
    dim_cols = [c for c in db_result.columns if c != target_col]

    # Fast path: an explicit single dimension_column with a single-part value.
    dim_col = query.get("dimension_column") if query else None
    if dim_col and len(parts) == 1:
        dim_actual = col_lower.get(str(dim_col).lower())
        if dim_actual:
            mask = db_result[dim_actual].apply(lambda v: _cell_matches(parts[0], v))
            hit  = db_result[mask]
            if not hit.empty:
                return hit[target_col].iloc[0]

    # General path: every part must match some dimension column in one row.
    for _, row in db_result.iterrows():
        if all(any(_cell_matches(part, row[c]) for c in dim_cols) for part in parts):
            return row[target_col]

    return None


def llm_match_fallback(unmatched_kpis: list, reference_queries: list, client) -> list:
    """
    For KPIs that rule-based matching couldn't resolve, ask the LLM to map them.
    Returns list of {kpi, query, col_alias} dicts.
    """
    if not unmatched_kpis or not reference_queries:
        return []

    kpi_names = [k["kpi_name"] for k in unmatched_kpis]
    query_list = [
        f"{i}: {q['query_name']} (patterns: {q.get('kpi_patterns', [])})"
        for i, q in enumerate(reference_queries)
    ]

    prompt = (
        "Map each KPI to the best reference query index.\n"
        f"KPIs: {kpi_names}\n"
        f"Queries:\n" + "\n".join(query_list) + "\n"
        'Return ONLY JSON: [{"kpi_name":"","query_index":0,"column_alias":"","row_key_value":null}]'
    )

    try:
        import json as _json
        from utils.sql_generator import _call_llm
        results = _call_llm(prompt, client, max_tokens=300)
        matched = []
        for item in results:
            idx = item.get("query_index", -1)
            if 0 <= idx < len(reference_queries):
                kpi_name = item.get("kpi_name", "")
                kpi_obj  = next((k for k in unmatched_kpis if k["kpi_name"] == kpi_name), None)
                if kpi_obj:
                    matched.append({
                        "kpi":       kpi_obj,
                        "query":     reference_queries[idx],
                        "col_alias": item.get("column_alias", ""),
                        "dim_value": item.get("row_key_value"),
                    })
        return matched
    except Exception:
        return []


# ── Report helpers ─────────────────────────────────────────────────────────────

def tables_to_kpis(table_results: list) -> list:
    """
    Flatten scraped visual tables into a KPI list for validation.
    Mirrors the Step-1 flattening so combination workbooks compare identically.

    Each table dict: {visual_title, headers|columns, rows, has_row_header}.
    Returns [{visual_name, kpi_name, value, value_type, visual_type}].
      - table w/ dimension → kpi_name = "Metric [DimValue]"
      - pure-metric card   → kpi_name = "Metric"
    """
    kpis = []
    seen = set()
    for tbl in table_results or []:
        cols           = tbl.get("headers") or tbl.get("columns", [])
        rows           = tbl.get("rows", [])
        visual_name    = tbl.get("visual_title", "Table")
        has_row_header = tbl.get("has_row_header", True)

        for row in rows:
            if not row:
                continue
            if has_row_header and len(cols) > 1:
                row_id = row[0] if row else ""
                for ci in range(1, len(cols)):
                    if ci < len(row) and row[ci]:
                        name = f"{cols[ci]} [{row_id}]"
                        if name not in seen:
                            seen.add(name)
                            kpis.append({"visual_name": visual_name, "kpi_name": name,
                                         "value": row[ci], "value_type": "number",
                                         "visual_type": "table"})
            elif not has_row_header and cols:
                for ci in range(len(cols)):
                    if ci < len(row) and row[ci]:
                        name = cols[ci]
                        if name not in seen:
                            seen.add(name)
                            kpis.append({"visual_name": visual_name, "kpi_name": name,
                                         "value": row[ci], "value_type": "number",
                                         "visual_type": "card"})
            else:
                name = cols[0] if cols else "Value"
                if name not in seen:
                    seen.add(name)
                    kpis.append({"visual_name": visual_name, "kpi_name": name,
                                 "value": row[0] if row else "", "value_type": "number",
                                 "visual_type": "table"})
    return kpis


def build_validation_df(results: list) -> pd.DataFrame:
    """
    Build a tidy DataFrame from the list of validation result dicts.
    Each dict must have: visual_name, kpi_name, ui_value, sql_query,
                         db_value, status, reason
    """
    rows = [
        {
            "Visual Name": r.get("visual_name", ""),
            "KPI Name":    r.get("kpi_name", ""),
            "UI Value":    r.get("ui_value", ""),
            "SQL Query":   r.get("sql_query", ""),
            "DB Value":    r.get("db_value", ""),
            "Status":      r.get("status", ""),
            "Reason":      r.get("reason", ""),
        }
        for r in results
    ]
    return pd.DataFrame(rows)


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Render the validation DataFrame to an in-memory Excel workbook with
    colour-coded rows (green = PASS, red = FAIL, yellow = ERROR).
    Returns the raw bytes ready for st.download_button.
    """
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PBI Validation Report"

    # ── Styles ────────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="E8650A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    pass_fill    = PatternFill("solid", fgColor="C6EFCE")
    fail_fill    = PatternFill("solid", fgColor="FFC7CE")
    error_fill   = PatternFill("solid", fgColor="FFEB9C")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_top     = Alignment(wrap_text=True, vertical="top")

    # ── Header row ────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = wrap_center
        cell.border    = border

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        status   = str(getattr(row, "Status", ""))
        row_fill = pass_fill if status == "PASS" else (
                   fail_fill if status == "FAIL" else error_fill)

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=str(value) if value is not None else "")
            cell.fill      = row_fill
            cell.alignment = wrap_top
            cell.border    = border

    # ── Column widths ─────────────────────────────────────────────────────
    target_widths = [22, 28, 14, 60, 14, 10, 45]
    for i, w in enumerate(target_widths, 1):
        if i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(output)
    return output.getvalue()


# ── Phase-1 extraction workbook (visual-wise sheets + slicer sheet) ─────────────

def _safe_sheet_name(name: str, used: set) -> str:
    """
    Turn a visual title into a valid, unique Excel sheet name:
      - strip Excel-illegal chars  [ ] : * ? / \\
      - cap at 31 chars
      - de-duplicate with a numeric suffix
    """
    clean = re.sub(r"[\[\]:*?/\\]", " ", str(name or "")).strip()
    clean = re.sub(r"\s+", " ", clean) or "Visual"
    clean = clean[:31]

    candidate = clean
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = clean[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def export_visuals_workbook(table_results: list, slicers: list = None) -> bytes:
    """
    Phase-1 extraction export.

    Builds an Excel workbook with:
      - one sheet per visual, in the visual's native table shape
        (headers row + data rows, exactly as scraped from 'Show as a table')
      - a leading 'Slicers' sheet listing each slicer's selected value

    Args:
        table_results — list of {visual_title, headers, rows} dicts from
                        try_show_as_table()
        slicers       — list of {slicer_name, selected_value} dicts from
                        extract_slicers_via_dom()

    Returns raw .xlsx bytes for st.download_button.
    """
    slicers = slicers or []
    output  = BytesIO()
    wb      = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    header_fill = PatternFill("solid", fgColor="E8650A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_top    = Alignment(wrap_text=True, vertical="top")

    used_names = set()

    # ── Slicers sheet (first) ────────────────────────────────────────────────
    ws = wb.create_sheet(_safe_sheet_name("Slicers", used_names))
    for col_idx, col_name in enumerate(["Slicer", "Selected Value"], 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill, cell.font, cell.alignment, cell.border = (
            header_fill, header_font, wrap_center, border)
    for row_idx, s in enumerate(slicers, 2):
        ws.cell(row=row_idx, column=1,
                value=str(s.get("slicer_name", ""))).border = border
        ws.cell(row=row_idx, column=2,
                value=str(s.get("selected_value", ""))).border = border
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A2"

    # ── One sheet per visual (native shape), de-duplicated ───────────────────
    # Two levels of dedup so every detail appears exactly once in the workbook:
    #   seen_sheet_sigs — skip a visual whose entire table (headers + rows) is
    #                     identical to one already written (duplicate scrape).
    #   seen_rows       — skip any individual row already recorded in an earlier
    #                     sheet, so a given detail never repeats across sheets.
    seen_sheet_sigs = set()
    seen_rows       = set()

    for vi, tbl in enumerate(table_results or [], 1):
        title      = tbl.get("visual_title") or f"Visual {vi}"
        headers    = tbl.get("headers") or tbl.get("columns") or []
        raw_rows   = tbl.get("rows") or []

        # Whole-sheet dedup — identical content already written?
        sheet_sig = (tuple(str(h) for h in headers),
                     tuple(tuple(str(c) for c in r) for r in raw_rows))
        if sheet_sig in seen_sheet_sigs:
            continue
        seen_sheet_sigs.add(sheet_sig)

        # Row-level dedup — keep only rows not seen in an earlier sheet
        rows = []
        for r in raw_rows:
            rkey = tuple(str(c).strip() for c in r)
            if rkey in seen_rows:
                continue
            seen_rows.add(rkey)
            rows.append(r)

        # Nothing unique left to write → don't emit an empty sheet
        if not rows and raw_rows:
            continue

        ws = wb.create_sheet(_safe_sheet_name(title, used_names))

        start_row = 1
        # Keep the full title as a caption row when it was truncated for the tab
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        start_row = 3

        if headers:
            for col_idx, col_name in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
                cell.fill, cell.font, cell.alignment, cell.border = (
                    header_fill, header_font, wrap_center, border)
            data_start = start_row + 1
        else:
            data_start = start_row

        for r_off, row in enumerate(rows):
            for c_off, val in enumerate(row, 1):
                cell = ws.cell(row=data_start + r_off, column=c_off,
                               value="" if val is None else str(val))
                cell.alignment = wrap_top
                cell.border    = border

        # Column widths — size to widest cell (headers + data), capped
        ncols = max([len(headers)] + [len(r) for r in rows] + [1])
        for c in range(1, ncols + 1):
            longest = len(str(headers[c - 1])) if c - 1 < len(headers) else 0
            for row in rows:
                if c - 1 < len(row) and row[c - 1] is not None:
                    longest = max(longest, len(str(row[c - 1])))
            ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 12), 55)

        if headers:
            ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

    # Guard: openpyxl requires at least one sheet
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    wb.save(output)
    return output.getvalue()
