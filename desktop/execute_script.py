import time
import os
import pytest
from openpyxl import Workbook
from pywinauto.application import Application
from pywinauto.keyboard import send_keys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


EXCEL_PATH = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
FILE_NAME = "Application_details.xlsx"
USERNAME = "Sathanantham.aru@tigeranayltics.com"
PASSWORD = "************"


# -------------------------------------------------
# STEP 1: Create Excel File (NO UI DEPENDENCY)
# -------------------------------------------------
def create_excel():

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "application_name"
    ws["B1"] = "application_url"
    ws["C1"] = "user_type"

    ws.append(["google", "www.google.com", "admin"])
    ws.append(["amazon", "www.amazon.com", "localuser"])

    file_path = os.path.abspath(FILE_NAME)
    wb.save(file_path)

    print("Excel created:", file_path)

    return file_path


# -------------------------------------------------
# STEP 2: Open Excel & Save to OneDrive
# -------------------------------------------------
def save_excel_to_onedrive(file_path):

    app = Application(backend="uia").start(EXCEL_PATH)
    time.sleep(5)

    excel = app.window(title_re=".*Excel")
    excel.wait("visible", timeout=20)

    excel.set_focus()
    time.sleep(2)

    # Open file (CTRL + O)
    send_keys("^o")
    time.sleep(2)

    send_keys(file_path)
    send_keys("{ENTER}")
    time.sleep(5)

    excel.set_focus()

    # Save As (ALT + F → A)
    send_keys("%F")
    time.sleep(1)

    send_keys("A")
    time.sleep(3)

    # Select OneDrive (since already logged in)
    try:
        excel.child_window(title_re=".*OneDrive.*").click_input()
        time.sleep(3)
    except:
        print("OneDrive auto selection skipped")

    # Enter file name
    send_keys("Application_details")
    time.sleep(1)

    send_keys("{ENTER}")
    time.sleep(5)

    print("Saved to OneDrive")

    excel.close()


# -------------------------------------------------
# STEP 3: Open Browser AFTER Excel
# -------------------------------------------------
@pytest.fixture(scope="module")
def driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")  # 🔑 prevents Skia/SharedImage GPU errors
    chrome_options.add_argument("--disable-software-rasterizer")
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--remote-allow-origins=*")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # chrome_options.binary_location = chromedriver_path
    # service = Service(executable_path=chromedriver_path)
    # service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    yield driver
    driver.quit()


# -------------------------------------------------
# STEP 4: OneDrive Web Flow
# -------------------------------------------------
def onedrive_download(driver):

    driver.get("https://onedrive.live.com/login")
    time.sleep(5)

    # Email
    driver.find_element(By.NAME, "loginfmt").send_keys(USERNAME)
    driver.find_element(By.ID, "idSIButton9").click()
    time.sleep(5)

    # Password
    driver.find_element(By.NAME, "passwd").send_keys(PASSWORD)
    driver.find_element(By.ID, "idSIButton9").click()
    time.sleep(10)

    # Stay signed in
    try:
        driver.find_element(By.ID, "idSIButton9").click()
    except:
        pass

    time.sleep(10)

    # Search file
    search = driver.find_element(By.XPATH, "//input[@placeholder='Search']")
    search.send_keys("Application_details")
    search.send_keys(Keys.ENTER)

    time.sleep(5)

    # Click file
    driver.find_element(By.XPATH, "//span[text()='Application_details']").click()
    time.sleep(3)

    # Download
    driver.find_element(By.XPATH, "//button[@name='Download']").click()

    time.sleep(10)


# -------------------------------------------------
# MAIN TEST
# -------------------------------------------------
def test_end_to_end(driver):

    # Step 1 → Excel create
    file_path = create_excel()

    # Step 2 → Save to OneDrive
    save_excel_to_onedrive(file_path)

    # Step 3 → THEN open browser
    onedrive_download(driver)